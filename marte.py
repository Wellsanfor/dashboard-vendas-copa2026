import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

# === CONFIGURAÇÕES INICIAIS ===
NUM_REGISTROS = 200
nome_arquivo = 'Dashboard_Copa_Premium.xlsx'

# Produtos por categoria
PRODUTOS = {
    'Vestuário': [('Camisa Seleção', 250.00), ('Jaqueta Brasil', 450.00)],
    'Acessórios': [('Chaveiro Copa', 15.00), ('Copo Térmico', 80.00)],
    'Equipamentos': [('Bola Al Rihla', 599.00), ('Luvas Goleiro', 120.00)],
    'Decoração': [('Bandeira Brasil G', 45.00), ('Vuvuzela', 20.00)]
}

REGIOES = ['Norte', 'Nordeste', 'Sul', 'Sudeste', 'Centro-Oeste']
VENDEDORES = ['Ana Silva', 'Carlos Lima', 'Maria Souza', 'João Santos']

# === GERAÇÃO DE DADOS (PANDAS) ===
dados = []
data_atual = datetime(2026, 6, 1)

for i in range(NUM_REGISTROS):
    data_venda = data_atual + timedelta(days=random.randint(0, 30))
    cat = random.choice(list(PRODUTOS.keys()))
    prod, preco = random.choice(PRODUTOS[cat])
    qtd = random.randint(1, 15)
    faturamento = qtd * preco
    lucro = faturamento * random.uniform(0.15, 0.35)
    
    dados.append({
        'Data': data_venda, 'Pedido': f'#C{1000+i}', 'Produto': prod,
        'Categoria': cat, 'Região': random.choice(REGIOES), 
        'Quantidade': qtd, 'Faturamento': faturamento, 'Lucro': lucro
    })

df = pd.DataFrame(dados)

# Tabelas auxiliares
res_regiao = df.groupby('Região')['Faturamento'].sum().reset_index()
res_produto = df.groupby('Produto')['Quantidade'].sum().sort_values(ascending=False).reset_index()
res_categoria = df.groupby('Categoria')['Faturamento'].sum().reset_index()
res_tempo = df.groupby('Data')['Faturamento'].sum().reset_index()

# === CRIAÇÃO DO WORKBOOK ===
wb = Workbook()
ws = wb.active
ws.title = "Dashboard"
ws.sheet_view.showGridLines = False

# --- ESTILOS (CORREÇÃO AQUI) ---
AZUL_DARK = PatternFill(start_color="002776", end_color="002776", fill_type="solid")
VERDE_BRASIL = PatternFill(start_color="009C3B", end_color="009C3B", fill_type="solid")
AMARELO_OURO = PatternFill(start_color="FFDF00", end_color="FFDF00", fill_type="solid")
CINZA_FUNDO = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
BRANCO = Font(color="FFFFFF", bold=True) # Variável que faltava!

# 1. CABEÇALHO
ws.merge_cells('A1:P2')
ws['A1'] = "🏆 DASHBOARD ESTRATÉGICO - COPA DO MUNDO 2026"
ws['A1'].font = Font(size=22, bold=True, color="FFFFFF")
ws['A1'].fill = AZUL_DARK
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

# 2. CARDS DE KPI
kpi_labels = [("FATURAMENTO TOTAL", "B4"), ("LUCRO LÍQUIDO", "F4"), ("TOTAL ITENS", "J4"), ("MARGEM MÉDIA", "N4")]
for text, cell in kpi_labels:
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=10)
    ws[cell].alignment = Alignment(horizontal='center')
    col_idx = ws[cell].column
    valor_cel = ws.cell(row=5, column=col_idx)
    valor_cel.font = Font(size=16, bold=True)
    valor_cel.alignment = Alignment(horizontal='center')
    if "TOTAL" in text:
        valor_cel.value = f"R$ {df['Faturamento'].sum():,.2f}"
    elif "LUCRO" in text:
        valor_cel.value = f"R$ {df['Lucro'].sum():,.2f}"
    elif "ITENS" in text:
        valor_cel.value = int(df['Quantidade'].sum())
    else:
        valor_cel.value = "28.5%"

# 3. ABA AUXILIAR
ws_aux = wb.create_sheet("Dados_Graficos")
for r in dataframe_to_rows(res_regiao, index=False): ws_aux.append(r)
ws_aux.append([])
for r in dataframe_to_rows(res_produto, index=False): ws_aux.append(r)
ws_aux.append([])
for r in dataframe_to_rows(res_categoria, index=False): ws_aux.append(r)
ws_aux.append([])
for r in dataframe_to_rows(res_tempo, index=False): ws_aux.append(r)

# 4. GRÁFICOS
# G1: Região
c1 = BarChart()
c1.title = "Faturamento por Região"
c1.add_data(Reference(ws_aux, min_col=2, min_row=1, max_row=5))
c1.set_categories(Reference(ws_aux, min_col=1, min_row=1, max_row=5))
ws.add_chart(c1, "A7")

# G2: Categorias
c2 = PieChart()
c2.title = "Mix de Categorias"
c2.add_data(Reference(ws_aux, min_col=2, min_row=13, max_row=16))
c2.set_categories(Reference(ws_aux, min_col=1, min_row=13, max_row=16))
ws.add_chart(c2, "I7")

# G3: Evolução
c3 = LineChart()
c3.title = "Evolução das Vendas"
c3.add_data(Reference(ws_aux, min_col=2, min_row=18, max_row=48))
c3.set_categories(Reference(ws_aux, min_col=1, min_row=18, max_row=48))
ws.add_chart(c3, "A22")

# G4: Top Produtos
c4 = BarChart()
c4.type = "bar"
c4.title = "Top 5 Produtos"
c4.add_data(Reference(ws_aux, min_col=2, min_row=7, max_row=11))
c4.set_categories(Reference(ws_aux, min_col=1, min_row=7, max_row=11))
ws.add_chart(c4, "I22")

# 5. ÁREA DE IMAGENS
ws.merge_cells('A38:P38')
ws['A38'] = "📸 GALERIA DE PRODUTOS (Arraste suas imagens para os quadros abaixo)"
ws['A38'].fill = VERDE_BRASIL
ws['A38'].font = BRANCO
ws['A38'].alignment = Alignment(horizontal='center')

quadros = ['B40:D45', 'F40:H45', 'J40:L45', 'N40:P45']
for q in quadros:
    ws.merge_cells(q)
    start_cell = q.split(':')[0]
    ws[start_cell].fill = CINZA_FUNDO
    ws[start_cell].alignment = Alignment(horizontal='center', vertical='center')
    ws[start_cell] = "[ FOTO ]"

# 6. ABA DE BASE DE DADOS
ws_base = wb.create_sheet("Base_Dados")
for r in dataframe_to_rows(df, index=False, header=True):
    ws_base.append(r)

# AJUSTE DE COLUNAS
for col in range(1, 17):
    ws.column_dimensions[get_column_letter(col)].width = 14

wb.save(nome_arquivo)
print(f"✅ Sucesso! Dashboard premium gerado em: {nome_arquivo}")